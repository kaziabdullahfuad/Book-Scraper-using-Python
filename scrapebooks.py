import os
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

url="http://books.toscrape.com/"

all_urls=[]

os.chdir("D:\\Python\\Web Scraping\\Books to Scrape")

excel_book=Workbook()
excel_sheet=excel_book.active

all_urls.append(url)

headers={
    "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36"
}

#req=requests.get(url,headers=headers)

#print(req.request.headers)

#soup=BeautifulSoup(req.content,features="lxml")

#print(soup.prettify())
# Need to find the name of the book,price and rating and in stock or not. So total-4 columns
excel_sheet.cell(row=1,column=1).value="Serial Number"
excel_sheet.cell(row=1,column=2).value="Book Name"
excel_sheet.cell(row=1,column=3).value="Book Price"
excel_sheet.cell(row=1,column=4).value="Book Rating"
excel_sheet.cell(row=1,column=5).value="Book Availability"

for i in range(2,6):
    extra=f"catalogue/page-{i}.html"
    all_urls.append(os.path.join(url,extra))

# for link in all_urls:
#     print(link)

#container=soup.find_all("li",{"class":"col-xs-6 col-sm-4 col-md-3 col-lg-3"}) # returns a list of all the book "li"
#print(container)
#print(len(container))
# book_name=container[0].find("img")['alt']
# book_rating=container[0].find("p")['class'][1]
# book_price=container[0].find("p",{"class":"price_color"}).text
# book_stock=container[0].find("p",{"class":"instock availability"}).text.strip()
# print(book_name)
# print(book_rating)
# print(book_price)
# print(book_stock)
count=0
sl=1
i=2
for link in all_urls:
    req=requests.get(link,headers=headers)
    soup=BeautifulSoup(req.content,features="lxml")
    container=soup.find_all("li",{"class":"col-xs-6 col-sm-4 col-md-3 col-lg-3"})

    for contain in container:
        book_name=contain.find("img")['alt']
        book_rating=contain.find("p")['class'][1]
        book_price=contain.find("p",{"class":"price_color"}).text
        book_stock=contain.find("p",{"class":"instock availability"}).text.strip() # gets rid of all the whitespaces before and after
        #print(book_name,book_rating,book_price,book_stock,end=" ")
        excel_sheet.cell(row=i,column=1).value=sl
        excel_sheet.cell(row=i,column=2).value=book_name
        excel_sheet.cell(row=i,column=3).value=book_price
        excel_sheet.cell(row=i,column=4).value=book_rating
        excel_sheet.cell(row=i,column=5).value=book_stock
        count+=1
        sl+=1
        i+=1
    
    print("\n")

print("Total: ",count)
excel_book.save("Books Scraped.xlsx")

# i=2
# extra=f"catalogue/page-{i}.html"
# print(os.path.join(url,extra))
# i+=1
# extra=f"catalogue/page-{i}.html"
# print(os.path.join(url,extra))